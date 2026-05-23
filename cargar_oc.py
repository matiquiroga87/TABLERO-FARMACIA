"""
cargar_oc.py — Matching con prioridad: solicitud+renglon > solicitud+codigo > codigo unico
"""
import openpyxl
from openpyxl import load_workbook
from typing import Optional

RUBROS_CFG = [
    ('ESTERILIZACION',3,53),('UTI',3,53),('ANESTESICOS',3,21),
    ('COMPRIMIDOS',3,80),('ANTIBIOTICOS',3,79),('PSICOTROPICOS',3,79),
    ('AMPOLLAS',3,106),('SUEROS',3,27),('DROGAS Y LIQUIDOS',3,38),
    ('ALIMENTACION',3,31),('SIES SAT URS',3,45),('BAJO COSTO',3,213),
    ('CREMAS GOTAS AEROSOLES',3,68),('DIALISIS',3,27),('RAYOS',3,33),
    ('FORMULACION Y SEDRONAR',3,32),('VARIOS',3,9),
    ('DESIERTOS',3,178),('ESTERILIZACION DESIERTOS',3,13),
]

# Índices de columnas (base 0 — openpyxl usa base 1 por eso sumamos 1 al escribir)
COL_CODIGO    = 0   # A — código vademecum
COL_SOLICITUD = 10  # K — número de solicitud
COL_PROVEEDOR = 12  # M — proveedor
COL_RENGLON   = 14  # O — número de renglón en la licitación
COL_OC        = 15  # P — N° OC emitida
COL_CANT_OC   = 20  # U — cantidad con OC 2026
COL_PRECIO    = 24  # Y — precio proveedor 2026
COL_MONTO     = 26  # AA — monto total adjudicado S/P proveedor


def _v(ws, row, col0):
    """Lee una celda (col en base 0) como string limpio."""
    val = ws.cell(row, col0 + 1).value
    return str(val).strip() if val is not None else ""


def construir_indice(wb: openpyxl.Workbook) -> dict:
    """
    Recorre todas las hojas UNA sola vez y construye tres índices:

    idx_sol_reng[(solicitud, renglon)] → (sheet, row)   ← más específico
    idx_sol_cod [(solicitud, codigo)]  → (sheet, row)
    idx_cod     [codigo]               → [(sheet, row), ...]  ← para detectar únicos
    """
    idx_sol_reng = {}
    idx_sol_cod  = {}
    idx_cod      = {}

    for sheet_name, r1, r2 in RUBROS_CFG:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in range(r1, r2 + 1):
            codigo    = _v(ws, row, COL_CODIGO).upper()
            solicitud = _v(ws, row, COL_SOLICITUD)
            renglon   = _v(ws, row, COL_RENGLON)

            if not codigo:
                continue

            # Índice 1: solicitud + renglón (combinación más específica)
            if solicitud and renglon:
                k = (solicitud, renglon)
                if k not in idx_sol_reng:
                    idx_sol_reng[k] = (sheet_name, row)

            # Índice 2: solicitud + código
            if solicitud:
                k2 = (solicitud, codigo)
                if k2 not in idx_sol_cod:
                    idx_sol_cod[k2] = (sheet_name, row)

            # Índice 3: código solo → lista para detectar si es único
            idx_cod.setdefault(codigo, []).append((sheet_name, row))

    return {"sol_reng": idx_sol_reng, "sol_cod": idx_sol_cod, "cod": idx_cod}


def buscar_renglon(indice: dict, codigo: str, nro_solicitud: str,
                   nro_renglon: str) -> tuple:
    """
    Devuelve (ubicacion, metodo) donde:
      ubicacion = (sheet_name, row) o None
      metodo    = string explicando cómo se encontró

    Prioridad:
      1. N° solicitud + N° renglón   ← más confiable, evita confusión entre solicitudes
      2. N° solicitud + código       ← segundo nivel
      3. Código único en todo el archivo
      4. No encontrado (código ambiguo o inexistente)
    """
    codigo    = codigo.strip().upper()
    solicitud = str(nro_solicitud).strip()
    renglon   = str(nro_renglon).strip()

    # 1. Solicitud + renglón
    if solicitud and renglon:
        k = (solicitud, renglon)
        if k in indice["sol_reng"]:
            return indice["sol_reng"][k], "Solicitud + Renglón"

    # 2. Solicitud + código
    if solicitud and codigo:
        k2 = (solicitud, codigo)
        if k2 in indice["sol_cod"]:
            return indice["sol_cod"][k2], "Solicitud + Código"

    # 3. Código único
    if codigo in indice["cod"]:
        apariciones = indice["cod"][codigo]
        if len(apariciones) == 1:
            return apariciones[0], "Código único"
        else:
            return None, f"Código ambiguo — aparece en {len(apariciones)} filas"

    return None, "No encontrado"


def aplicar_oc(ruta_xlsx: str, datos_oc: dict,
               renglones_confirmados: list) -> tuple:
    """
    Aplica los renglones confirmados al workbook.
    Devuelve (workbook_modificado, lista_de_resultados).
    """
    wb       = load_workbook(ruta_xlsx)
    indice   = construir_indice(wb)
    nro_oc   = datos_oc["encabezado"].get("oc_completo", "")
    solicitud= datos_oc["encabezado"].get("nro_solicitud", "")
    proveedor= datos_oc["proveedor"].get("razon_social", "")
    resultados = []

    for reng in renglones_confirmados:
        codigo   = str(reng.get("codigo", "")).strip().upper()
        nro_reng = str(reng.get("renglon", "")).strip()
        cantidad = reng.get("cantidad_num", 0)
        precio   = reng.get("importe_unitario_num", 0)
        monto    = reng.get("importe_total_num", 0)

        ubicacion, metodo = buscar_renglon(indice, codigo, solicitud, nro_reng)

        if ubicacion:
            sheet_name, row = ubicacion
            ws = wb[sheet_name]
            ws.cell(row, COL_OC      + 1).value = nro_oc
            ws.cell(row, COL_CANT_OC + 1).value = cantidad
            ws.cell(row, COL_PRECIO  + 1).value = precio
            ws.cell(row, COL_MONTO   + 1).value = monto
            # Proveedor: solo si estaba vacío
            prov_actual = ws.cell(row, COL_PROVEEDOR + 1).value
            if not prov_actual or str(prov_actual).strip().lower() in ["", "nan"]:
                ws.cell(row, COL_PROVEEDOR + 1).value = proveedor
            resultados.append({
                "renglon": nro_reng, "codigo": codigo,
                "descripcion": reng.get("descripcion","")[:60],
                "match": metodo, "estado": "✅ Cargado",
                "hoja": sheet_name, "fila": row,
            })
        else:
            resultados.append({
                "renglon": nro_reng, "codigo": codigo,
                "descripcion": reng.get("descripcion","")[:60],
                "match": metodo, "estado": "⚠️ No encontrado",
                "hoja": "—", "fila": "—",
            })

    return wb, resultados
